from openpyxl.workbook import workbook
from openpyxl import load_workbook 

wb = load_workbook('Employeedata.xlsx') #am using a library to open the excel file
ws = wb.active 
sheet = wb['Sheet1']


for i in range( 2, sheet.max_row+1):
     cell = sheet.cell(i, 2)
     if 'helpinghands.cm' in cell.value: 
          update = (cell.value).replace('helpinghands.cm', 'handsinhands.org') # chnages from xlsx to .org
          sheet.cell(i, 2).value = update
wb.save('updated_emails.xlsx') # save files